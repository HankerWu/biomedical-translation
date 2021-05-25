import argparse
import config
from packaging.version import parse
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM, DataCollatorForSeq2Seq, Seq2SeqTrainingArguments, Seq2SeqTrainer
from datasets import load_metric
from nltk.data import load
from get_pdf_content import PDF_Content
import pandas as pd
import os

from nltk.tokenize.punkt import PunktSentenceTokenizer, PunktParameters

nltk_tokenizer = load("tokenizers/punkt/{0}.pickle".format("english"))
nltk_tokenizer._params.abbrev_types.update({'al', 'e.g', 'resp', 'i.e'})
use_args=False

if use_args:
    # 创建解析步骤
    parser = argparse.ArgumentParser()

    # 添加参数步骤
    parser.add_argument('--pdf',
                        action="store_true",
                        help='whether input is pdf or not')
    parser.add_argument('--field',
                        choices=['abstract', 'title', 'introduction','all'],
                        help='translation field')
    parser.add_argument('--filename',
                        help='name of file')

    # 解析参数步骤
    args = parser.parse_args()
    print("argparse.args=", args, type(args))

tokenizer = AutoTokenizer.from_pretrained('D:/MSRA/project/transformers_model/best')
metric = load_metric("metric.py")

model = AutoModelForSeq2SeqLM.from_pretrained(config.model_dir)
text = []

if use_args:
    if args.pdf:
        print('pdf mode.')
        pdf_content = PDF_Content(args.filename)
        pdf_content.getContent()
        if args.field == 'title':
            text = pdf_content.title
        elif args.field == 'abstract':
            text = pdf_content.abstract.split('\n')[1:]
        elif args.field == 'introduction':
            text = pdf_content.introduction.split('\n')[1:]
        else:
            text += [pdf_content.title + '\n']
            text += pdf_content.abstract.split('\n')
            text += pdf_content.introduction.split('\n')
    else:
        print('txt mode.')
        # with open(args.filename, 'r') as f:
        #     text = f.read()
else:
    data_dir = 'D:/MSRA/project/get_data/test_article/'
    files = os.listdir(data_dir)
    for file in files:
        pdf_content = PDF_Content(data_dir+file)
        pdf_content.getContent()
        text += [pdf_content.title + '\n']
        text += pdf_content.abstract.split('\n')
        text += pdf_content.introduction.split('\n')
en_sent=[]
for t in text:
    en_sent += nltk_tokenizer.tokenize(t)
print('\n'.join(en_sent))

# model.to('cuda')
text = [
    "A  framework  to  identify  antigen-expanded  T  Cell  Receptor  (TCR)  clusters  within  complex repertoires"
]
# tokenized_text = tokenizer([text], return_tensors='pt').to('cuda')
batchsize = 20
sentnum = len(en_sent)
batchnum = int(sentnum / batchsize)

translated_text = []
for i in range(batchnum):
    tokenized_text = tokenizer(en_sent[i * 20:max(i * 20 + 20, sentnum)],
                               return_tensors='pt',
                               padding=True,
                               truncation=True)
    translation = model.generate(**tokenized_text)
    translated_text += tokenizer.batch_decode(translation,
                                            skip_special_tokens=True)
print('\n'.join(translated_text))

data_dict = {'source text': en_sent, 'translated text': translated_text}
df = pd.DataFrame(data_dict)

import openpyxl
import numpy as np

def write2xlsx(path, sheetname, value):
    index = len(value)
    workbook = openpyxl.Workbook()  # 实例化
    sheet = workbook.active  # 激活worksheet
    sheet.title = sheetname
    for i in range(index):
        for j in range(len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    # print("xlsx格式表格数据写入成功！")

def append2xlsx(path, sheetname, value):
    index = len(value)
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]

    for i in range(index):
        sheet.append(list(value[i]))  # append的内容必须是可迭代对象，里面的value必须是str类型
    workbook.save(path)
    # print("xlsx格式表格数据追加成功！")

column_name = np.array([df.columns])
write2xlsx(path='./data.xlsx', sheetname='text', value=column_name)

data = df.to_numpy()
append2xlsx(path='./data.xlsx', sheetname='text', value=data)
